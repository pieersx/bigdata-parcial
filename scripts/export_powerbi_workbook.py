import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pyspark.sql import SparkSession
from pyspark.sql import functions as F


ROOT = Path("/home/jovyan/work")
GOLD = ROOT / "data" / "gold"
OUTPUT = ROOT / "data" / "powerbi" / "powerbi_municipal_gold.xlsx"
HTML_OUTPUT = ROOT / "data" / "powerbi" / "dashboard_municipal_gold.html"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def read_gold(spark, table_name):
    return spark.read.parquet(str(GOLD / table_name))


def save_workbook_with_fallback(workbook, output_path):
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def add_sheet(workbook, title, dataframe):
    worksheet = workbook.create_sheet(title)
    columns = dataframe.columns
    worksheet.append(columns)
    for row in dataframe.toLocalIterator():
        worksheet.append([row[column] for column in columns])

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if worksheet.max_row > 1:
        table_name = "tbl_" + "".join(char if char.isalnum() else "_" for char in title)
        table = Table(displayName=table_name[:250], ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for index, column in enumerate(columns, start=1):
        max_length = max(len(str(column)), 12)
        for row in worksheet.iter_rows(min_row=2, max_row=min(worksheet.max_row, 150), min_col=index, max_col=index):
            if row[0].value is not None:
                max_length = max(max_length, min(len(str(row[0].value)), 36))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 38)


def collect_rows(dataframe, limit=5000):
    return [row.asDict(recursive=True) for row in dataframe.limit(limit).collect()]


def write_dashboard_html(payload):
    HTML_OUTPUT.write_text(
        f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Dashboard Municipal Gold</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <style>
    :root {{
      --green: #16a52e;
      --green2: #45b80f;
      --pale: #ddffd8;
      --panel: #fbfffb;
      --text: #242424;
      --muted: #686868;
      --orange: #ff8a1c;
      --teal: #10b8ac;
      --red: #e85d75;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Segoe UI, Arial, sans-serif; color: var(--text); background: #eefee9; }}
    .page {{ min-height: 100vh; padding: 0 42px 36px; background: var(--pale); display: none; }}
    .page.active {{ display: block; }}
    header {{ background: var(--green); color: white; text-align: center; font-size: 24px; font-weight: 700; padding: 14px 20px; }}
    .top {{ display: grid; grid-template-columns: 560px repeat(3, 1fr); gap: 20px; align-items: center; padding: 28px 36px 18px; border-bottom: 3px solid var(--green); }}
    .years, .cats {{ display: flex; flex-wrap: wrap; gap: 2px; }}
    button {{ border: 1px solid #b7dcb1; background: #eee; padding: 14px 22px; font-size: 16px; cursor: pointer; }}
    button.active {{ background: #303030; color: white; }}
    .kpi {{ text-align: center; }}
    .kpi .label {{ font-size: 22px; }}
    .kpi .value {{ font-size: 42px; line-height: 1.1; }}
    .body {{ display: grid; grid-template-columns: 120px 1fr 1fr; gap: 22px; padding-top: 34px; }}
    .side-title {{ font-size: 20px; margin-bottom: 8px; }}
    .cat-btn {{ width: 90px; height: 94px; display: block; margin-bottom: 6px; border-top: 2px solid #666; border-bottom: 2px solid #666; }}
    .panel {{ background: var(--panel); padding: 10px 16px 16px; min-height: 330px; position: relative; }}
    .panel h2 {{ margin: 0 0 10px; font-size: 22px; font-weight: 500; }}
    .wide {{ grid-column: span 2; }}
    canvas {{ width: 100% !important; max-height: 360px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th {{ background: var(--green2); color: white; text-align: left; padding: 8px; }}
    td {{ border-bottom: 1px solid #ddd; padding: 7px; }}
    .footer {{ padding: 16px 30px 0; color: #006b4f; font-size: 12px; }}
    .nav {{ position: fixed; bottom: 12px; left: 50%; transform: translateX(-50%); display: flex; gap: 6px; background: rgba(255,255,255,.92); padding: 6px 10px; border-radius: 18px; box-shadow: 0 2px 10px #0002; }}
    .nav button {{ padding: 6px 10px; font-size: 13px; }}
    .note {{ font-size: 15px; color: var(--muted); }}
  </style>
</head>
<body>
  <div id="app"></div>
  <nav class="nav" id="nav"></nav>
  <script>
    const DATA = {json.dumps(payload, ensure_ascii=False, default=str)};
  </script>
  <script>
    const fmt = new Intl.NumberFormat('es-PE', {{ maximumFractionDigits: 0 }});
    const fmt1 = new Intl.NumberFormat('es-PE', {{ maximumFractionDigits: 1 }});
    const years = [...new Set(DATA.d1.map(d => d.ANO_DOC))].sort((a,b)=>a-b);
    const cats = ['A','B','C','D','E','F','G','Sin categoria'];
    let state = {{ page: 0, year: years[years.length - 1], cat: 'Todas' }};
    let charts = [];

    const pages = [
      'Recaudación Municipal Vs Capacidad Tributaria',
      'Recaudación Por Clasificador De Ingreso',
      'Predial Vs Efectividad',
      'Distribución De Efectividad Predial',
      'Software Tributario Municipal',
      'Priorización De Municipalidades'
    ];

    function n(v) {{ return Number(v || 0); }}
    function catOk(d) {{ return state.cat === 'Todas' || (d.categoria_municipalidad || 'Sin categoria') === state.cat; }}
    function byYear(rows, field='ANO_DOC') {{ return rows.filter(d => Number(d[field]) === Number(state.year) && catOk(d)); }}
    function sum(rows, field) {{ return rows.reduce((a,d)=>a+n(d[field]),0); }}
    function top(rows, label, value, count=18) {{
      const m = new Map();
      rows.forEach(d => m.set(d[label] || 'SIN DATO', (m.get(d[label] || 'SIN DATO') || 0) + n(d[value])));
      return [...m.entries()].sort((a,b)=>b[1]-a[1]).slice(0,count);
    }}
    function renderButtons() {{
      return `<div class="years">${{years.map(y=>`<button onclick="state.year=${{y}}; render()" class="${{state.year==y?'active':''}}">${{y}}</button>`).join('')}}</div>`;
    }}
    function renderCats() {{
      return `<div><div class="side-title">Clasificación</div><button class="cat-btn ${{state.cat==='Todas'?'active':''}}" onclick="state.cat='Todas'; render()">Todas</button>${{cats.map(c=>`<button class="cat-btn ${{state.cat===c?'active':''}}" onclick="state.cat='${{c}}'; render()">${{c}}</button>`).join('')}}</div>`;
    }}
    function kpi(label, value) {{ return `<div class="kpi"><div class="label">${{label}}</div><div class="value">${{value}}</div></div>`; }}
    function panel(id, title) {{ return `<section class="panel"><h2>${{title}}</h2><canvas id="${{id}}"></canvas></section>`; }}
    function tablePanel(title, rows, cols) {{
      return `<section class="panel"><h2>${{title}}</h2><table><thead><tr>${{cols.map(c=>`<th>${{c[0]}}</th>`).join('')}}</tr></thead><tbody>${{rows.slice(0,24).map(r=>`<tr>${{cols.map(c=>`<td>${{c[2] ? c[2](r[c[1]]) : (r[c[1]] ?? '')}}</td>`).join('')}}</tr>`).join('')}}</tbody></table></section>`;
    }}
    function mkBar(id, rows, color=DATA.colors.green2) {{
      charts.push(new Chart(document.getElementById(id), {{ type:'bar', data: {{ labels: rows.map(r=>r[0]), datasets:[{{ data:rows.map(r=>r[1]), backgroundColor:color }}] }}, options: {{ indexAxis:'y', plugins:{{legend:{{display:false}}}}, scales:{{x:{{grid:{{color:'#d5ecd0'}}}},y:{{grid:{{display:false}}}}}} }} }}));
    }}
    function mkLine(id, rows) {{
      charts.push(new Chart(document.getElementById(id), {{ type:'line', data: {{ labels: rows.map(r=>r.periodo_id), datasets:[
        {{ label:'PIA', data:rows.map(r=>n(r.MONTO_PIA)), borderColor:'#86c65b', tension:.25 }},
        {{ label:'PIM', data:rows.map(r=>n(r.MONTO_PIM)), borderColor:'#0a7f22', tension:.25 }},
        {{ label:'Recaudado', data:rows.map(r=>n(r.MONTO_RECAUDADO)), borderColor:'#ff8a1c', tension:.25 }}
      ] }}, options: {{ plugins:{{legend:{{position:'top'}}}} }} }}));
    }}
    function mkScatter(id, rows, x, y, label) {{
      charts.push(new Chart(document.getElementById(id), {{ type:'scatter', data: {{ datasets:[{{ label, data:rows.map(r=>({{x:n(r[x]), y:n(r[y])}})), backgroundColor:'#ff8a1c' }}] }}, options: {{ plugins:{{legend:{{display:false}}}}, scales:{{x:{{title:{{display:true,text:x}}}}, y:{{title:{{display:true,text:y}}}}}} }} }}));
    }}
    function mkHist(id, rows, field) {{
      const vals = rows.map(r=>n(r[field])).filter(v=>v>=0);
      const bins = Array(10).fill(0);
      vals.forEach(v=>bins[Math.min(9, Math.floor(v/10))]++);
      charts.push(new Chart(document.getElementById(id), {{ type:'bar', data: {{ labels:bins.map((_,i)=>`${{i*10}}-${{i*10+10}}%`), datasets:[{{data:bins, backgroundColor:'#10b8ac'}}] }}, options:{{plugins:{{legend:{{display:false}}}}}} }}));
    }}
    function pageShell(title, kpis, body) {{
      return `<section class="page active"><header>Perú. ${{title}}</header><div class="top">${{renderButtons()}}${{kpis.join('')}}</div><main class="body">${{renderCats()}}${{body}}</main><div class="footer">Fuente: SIAF, SISMEPRE, RENAMU y CategoriasMunicipalidades / Elaboración propia</div></section>`;
    }}
    function render() {{
      charts.forEach(c=>c.destroy()); charts = [];
      const app = document.getElementById('app');
      const nav = document.getElementById('nav');
      nav.innerHTML = pages.map((p,i)=>`<button class="${{state.page===i?'active':''}}" onclick="state.page=${{i}}; render()">${{i+1}}</button>`).join('');
      let html = '';
      setTimeout(draw, 0);
      if (state.page === 0) {{
        const rows = byYear(DATA.mart);
        html = pageShell(pages[0], [
          kpi('Recaudación', fmt.format(sum(rows,'MONTO_RECAUDADO'))),
          kpi('Personal municipal', fmt.format(sum(rows,'personal_municipal_total'))),
          kpi('Recaud. media por personal', fmt1.format(sum(rows,'MONTO_RECAUDADO') / Math.max(1,sum(rows,'personal_municipal_total'))))
        ], panel('c1','Recaudación según departamento') + panel('c2','Recaudación vs personal municipal'));
      }} else if (state.page === 1) {{
        const rows = byYear(DATA.clasificador);
        html = pageShell(pages[1], [
          kpi('Recaudación', fmt.format(sum(rows,'MONTO_RECAUDADO'))),
          kpi('PIM', fmt.format(sum(rows,'MONTO_PIM'))),
          kpi('PIA', fmt.format(sum(rows,'MONTO_PIA')))
        ], panel('c1','Recaudación según clasificador') + tablePanel('Detalle territorial', byYear(DATA.avance), [['Departamento','DEPARTAMENTO_NOMBRE'],['Provincia','PROVINCIA_NOMBRE'],['Distrito','DISTRITO_NOMBRE'],['Ingreso','MONTO_RECAUDADO', v=>fmt.format(n(v))]]));
      }} else if (state.page === 2) {{
        const rows = byYear(DATA.predial, 'ANO_ESTADISTICA');
        html = pageShell(pages[2], [
          kpi('Predial', fmt.format(sum(rows,'MON_RECAUDACION_TOTAL'))),
          kpi('Saldo', fmt.format(sum(rows,'MON_SALDO_PREDIAL_TOTAL'))),
          kpi('Contribuyentes', fmt.format(sum(rows,'NUM_CONTRIPREDIO')))
        ], panel('c1','Predial por departamento') + panel('c2','Recaudación predial vs saldo'));
      }} else if (state.page === 3) {{
        const rows = byYear(DATA.priorizacion, 'ANO_ESTADISTICA');
        html = pageShell(pages[3], [
          kpi('Recuperación media', fmt1.format(sum(rows,'PCT_RECUPERACION_PREDIAL') / Math.max(1, rows.length)) + '%'),
          kpi('Municipalidades', fmt.format(new Set(rows.map(r=>r.SEC_EJEC)).size)),
          kpi('Categoría', state.cat)
        ], panel('c1','Efectividad predial según departamento') + panel('c2','Distribución de recuperación predial'));
      }} else if (state.page === 4) {{
        const rows = DATA.software.filter(catOk);
        html = pageShell(pages[4], [
          kpi('SRTM Estado', fmt.format(rows.filter(r=>r.usa_srtm_estado).length)),
          kpi('Software rentas', fmt.format(rows.filter(r=>r.usa_software_rentas_at).length)),
          kpi('Software catastro', fmt.format(rows.filter(r=>r.usa_software_catastro).length))
        ], panel('c1','Uso de SRTM por departamento') + panel('c2','Software de rentas y catastro'));
      }} else {{
        const rows = byYear(DATA.mart).sort((a,b)=>n(b.MON_SALDO_PREDIAL_TOTAL)-n(a.MON_SALDO_PREDIAL_TOTAL));
        html = pageShell(pages[5], [
          kpi('Saldo predial', fmt.format(sum(rows,'MON_SALDO_PREDIAL_TOTAL'))),
          kpi('Recaudación', fmt.format(sum(rows,'MONTO_RECAUDADO'))),
          kpi('Municipalidades', fmt.format(new Set(rows.map(r=>r.SEC_EJEC)).size))
        ], tablePanel('Cuadro de priorización municipal', rows, [['Departamento','DEPARTAMENTO_NOMBRE'],['Provincia','PROVINCIA_NOMBRE'],['Distrito','DISTRITO_NOMBRE'],['Categoría','categoria_municipalidad'],['Recaudación','MONTO_RECAUDADO', v=>fmt.format(n(v))],['Saldo predial','MON_SALDO_PREDIAL_TOTAL', v=>fmt.format(n(v))],['Recuperación','PCT_RECUPERACION_PREDIAL', v=>fmt1.format(n(v))+'%'],['SRTM','usa_srtm_estado', v=>v?'Sí':'No']]) + '<section class="panel"><h2>Criterios</h2><p class="note">Priorizar municipalidades con alto saldo predial, baja recuperación, sin software tributario y cumplimiento SISMEPRE débil.</p></section>');
      }}
      app.innerHTML = html;
    }}
    function draw() {{
      if (!document.getElementById('c1')) return;
      if (state.page === 0) {{
        const rows = byYear(DATA.mart);
        mkBar('c1', top(rows,'DEPARTAMENTO_NOMBRE','MONTO_RECAUDADO'), '#45b80f');
        mkScatter('c2', rows.slice(0,2000), 'personal_municipal_total', 'MONTO_RECAUDADO', 'Municipalidades');
      }} else if (state.page === 1) {{
        mkBar('c1', top(byYear(DATA.clasificador),'ESPECIFICA_DET_NOMBRE','MONTO_RECAUDADO'), '#45b80f');
      }} else if (state.page === 2) {{
        const rows = byYear(DATA.predial,'ANO_ESTADISTICA');
        mkBar('c1', top(rows,'DEPARTAMENTO_NOMBRE','MON_RECAUDACION_TOTAL'), '#45b80f');
        mkScatter('c2', rows.slice(0,2000), 'MON_SALDO_PREDIAL_TOTAL', 'MON_RECAUDACION_TOTAL', 'Predial');
      }} else if (state.page === 3) {{
        const rows = byYear(DATA.priorizacion,'ANO_ESTADISTICA');
        mkBar('c1', top(rows,'DEPARTAMENTO_NOMBRE','PCT_RECUPERACION_PREDIAL'), '#45b80f');
        mkHist('c2', rows, 'PCT_RECUPERACION_PREDIAL');
      }} else if (state.page === 4) {{
        const rows = DATA.software.filter(catOk);
        mkBar('c1', top(rows.filter(r=>r.usa_srtm_estado),'DEPARTAMENTO_NOMBRE','uno'), '#45b80f');
        const m = top(rows.filter(r=>r.usa_software_rentas_at),'DEPARTAMENTO_NOMBRE','uno').map(([k,v])=>[k, v + top(rows.filter(r=>r.usa_software_catastro),'DEPARTAMENTO_NOMBRE','uno').find(x=>x[0]===k)?.[1] || v]);
        mkBar('c2', m, '#f1a383');
      }}
    }}
    DATA.software.forEach(r=>r.uno=1);
    render();
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )


def main():
    spark = SparkSession.builder.appName("powerbi-municipal-workbook").getOrCreate()
    spark.sparkContext.setLogLevel("WARN")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    municipios_source = read_gold(spark, "dim_municipalidad_gold")
    if "in_scope_presentacion" not in municipios_source.columns:
        municipios_source = municipios_source.withColumn("in_scope_presentacion", F.lit(True))
    if "categoria_municipalidad" not in municipios_source.columns:
        municipios_source = municipios_source.withColumn("categoria_municipalidad", F.lit(None).cast("string"))
    if "categoria_match_status" not in municipios_source.columns:
        municipios_source = municipios_source.withColumn("categoria_match_status", F.lit("missing_category_source"))
    municipios = municipios_source.select(
        "SEC_EJEC", "UBIGEO", "MUNICIPALIDAD_NOMBRE", "DEPARTAMENTO_NOMBRE",
        "PROVINCIA_NOMBRE", "DISTRITO_NOMBRE", "has_siaf", "has_sismepre", "renamu_match",
        "in_scope_presentacion", "categoria_municipalidad", "categoria_match_status",
    )
    municipios_scope = municipios.filter("in_scope_presentacion")
    ingresos = read_gold(spark, "fact_ingresos_mensuales")
    tiempo = read_gold(spark, "dim_tiempo")
    ubigeo = read_gold(spark, "dim_ubigeo")
    clasificador_dim = read_gold(spark, "dim_clasificador_ingreso")
    estado_sismepre = read_gold(spark, "dim_estado_sismepre")
    formulario_sismepre = read_gold(spark, "dim_formulario_sismepre")
    pregunta_sismepre = read_gold(spark, "dim_pregunta_sismepre")
    predial = read_gold(spark, "fact_predial_mensual")
    cumplimiento = read_gold(spark, "fact_sismepre_cumplimiento")
    calidad = read_gold(spark, "fact_calidad_datos")
    gestion = read_gold(spark, "fact_renamu_gestion_tributaria")
    software = read_gold(spark, "fact_renamu_software_at")
    mart = read_gold(spark, "mart_dashboard_municipal")

    evolucion = (
        ingresos.groupBy("ANO_DOC", "MES_DOC", "periodo_id")
        .agg(
            F.sum("MONTO_PIA").alias("MONTO_PIA"),
            F.sum("MONTO_PIM").alias("MONTO_PIM"),
            F.sum("MONTO_RECAUDADO").alias("MONTO_RECAUDADO"),
            F.sum("variacion_pim_pia").alias("VARIACION_PIM_PIA"),
        )
        .withColumn(
            "PCT_EJECUCION",
            F.when(F.col("MONTO_PIM") != 0, F.round(F.col("MONTO_RECAUDADO") / F.col("MONTO_PIM") * 100, 4)),
        )
        .orderBy("ANO_DOC", "MES_DOC")
    )
    avance = (
        ingresos.groupBy("SEC_EJEC", "ANO_DOC")
        .agg(
            F.sum("MONTO_PIA").alias("MONTO_PIA"),
            F.sum("MONTO_PIM").alias("MONTO_PIM"),
            F.sum("MONTO_RECAUDADO").alias("MONTO_RECAUDADO"),
        )
        .withColumn(
            "PCT_EJECUCION",
            F.when(F.col("MONTO_PIM") != 0, F.round(F.col("MONTO_RECAUDADO") / F.col("MONTO_PIM") * 100, 4)),
        )
        .join(municipios_scope, "SEC_EJEC", "left")
        .orderBy(F.desc("ANO_DOC"), F.desc("MONTO_RECAUDADO"))
    )
    ranking = (
        avance.groupBy(
            "ANO_DOC", "DEPARTAMENTO_NOMBRE", "PROVINCIA_NOMBRE",
            "categoria_municipalidad", "categoria_match_status", "in_scope_presentacion",
        )
        .agg(
            F.sum("MONTO_PIM").alias("MONTO_PIM"),
            F.sum("MONTO_RECAUDADO").alias("MONTO_RECAUDADO"),
            F.countDistinct("SEC_EJEC").alias("MUNICIPALIDADES"),
        )
        .withColumn(
            "PCT_EJECUCION",
            F.when(F.col("MONTO_PIM") != 0, F.round(F.col("MONTO_RECAUDADO") / F.col("MONTO_PIM") * 100, 4)),
        )
        .orderBy(F.desc("ANO_DOC"), F.desc("MONTO_RECAUDADO"))
    )
    clasificador_fact = read_gold(spark, "fact_ingresos_clasificador")
    clasificador_view = (
        clasificador_fact.join(
            municipios_scope.select(
                "SEC_EJEC", "categoria_municipalidad", "categoria_match_status", "in_scope_presentacion"
            ),
            "SEC_EJEC",
            "left",
        )
        .groupBy(
            "ANO_DOC", "clasificador_id",
            "categoria_municipalidad", "categoria_match_status", "in_scope_presentacion",
        )
        .agg(
            F.sum("MONTO_PIA").alias("MONTO_PIA"),
            F.sum("MONTO_PIM").alias("MONTO_PIM"),
            F.sum("MONTO_RECAUDADO").alias("MONTO_RECAUDADO"),
        )
        .join(clasificador_dim, "clasificador_id", "left")
        .withColumn(
            "PCT_EJECUCION",
            F.when(F.col("MONTO_PIM") != 0, F.round(F.col("MONTO_RECAUDADO") / F.col("MONTO_PIM") * 100, 4)),
        )
        .orderBy(F.desc("ANO_DOC"), F.desc("MONTO_RECAUDADO"))
    )
    predial_view = (
        predial.select(
            "SEC_EJEC", "ANO_ESTADISTICA", "MES_ESTADISTICA", "periodo_id",
            "MON_RECAUDACTUAL_ORDIN", "MON_RECAUDACTUAL_COAC", "MON_RECAUDACION_TOTAL",
            "MON_SALDO_PREDIAL_TOTAL", "MON_BASEIMPONIBLE_AFECTO", "NUM_CONTRIPREDIO",
            "NUM_PREDIOTOTAL", "TIPO_META",
        )
        .join(municipios_scope, "SEC_EJEC", "left")
        .orderBy("ANO_ESTADISTICA", "MES_ESTADISTICA")
    )
    sismepre_view = (
        cumplimiento.select(
            "SEC_EJEC", "ANO_APLICACION", "PERIODO", "ESTADO", "CLASIFICACION",
            "TIPO_META", "ORIGEN_INFORMACION", "IND_RESOL_ALCAL_ADJUNTO", "has_sismepre",
        )
        .join(municipios_scope.drop("has_sismepre"), "SEC_EJEC", "left")
        .orderBy("ANO_APLICACION", "PERIODO")
    )
    priorizacion_view = (
        predial.groupBy("SEC_EJEC", "ANO_ESTADISTICA")
        .agg(
            F.sum("MON_RECAUDACION_TOTAL").alias("MON_RECAUDACION_TOTAL"),
            F.sum("MON_SALDO_PREDIAL_TOTAL").alias("MON_SALDO_PREDIAL_TOTAL"),
            F.sum("MON_BASEIMPONIBLE_AFECTO").alias("MON_BASEIMPONIBLE_AFECTO"),
        )
        .withColumn(
            "PCT_RECUPERACION_PREDIAL",
            F.when(
                F.col("MON_RECAUDACION_TOTAL") + F.col("MON_SALDO_PREDIAL_TOTAL") != 0,
                F.round(
                    F.col("MON_RECAUDACION_TOTAL")
                    / (F.col("MON_RECAUDACION_TOTAL") + F.col("MON_SALDO_PREDIAL_TOTAL"))
                    * 100,
                    4,
                ),
            ),
        )
        .join(municipios_scope, "SEC_EJEC", "left")
        .orderBy(F.desc("ANO_ESTADISTICA"), F.desc("MON_SALDO_PREDIAL_TOTAL"))
    )
    gestion_view = (
        gestion.join(municipios_scope, "SEC_EJEC", "left")
        .orderBy("ANO_RENAMU", "DEPARTAMENTO_NOMBRE", "PROVINCIA_NOMBRE", "DISTRITO_NOMBRE")
    )
    software_view = (
        software.join(municipios_scope, "SEC_EJEC", "left")
        .orderBy("ANO_RENAMU", "DEPARTAMENTO_NOMBRE", "PROVINCIA_NOMBRE", "DISTRITO_NOMBRE")
    )
    mart_view = mart.orderBy(F.desc("ANO_DOC"), F.desc("MONTO_RECAUDADO"))
    calidad_view = calidad.select(
        "layer", "dataset", "check_type", "status", "timestamp",
        "records_checked", "records_passed", "records_failed", "failure_rate",
    ).orderBy(F.desc("timestamp"))

    dashboard_design = [
        [
            "01 Evolucion mensual",
            "Como evoluciona el PIA, PIM y recaudado municipal?",
            "Recaudacion total; PIM total; pct ejecucion; variacion PIM-PIA",
            "Cards KPI, linea mensual PIA/PIM/Recaudado, barras por anio, tabla top meses",
            "d1_evolucion",
            "Anio, mes",
        ],
        [
            "02 Avance municipal",
            "Que municipalidades ejecutan mejor su presupuesto de ingresos?",
            "Recaudacion; PIM; pct ejecucion; municipalidades",
            "Ranking horizontal por municipalidad, scatter PIM vs recaudado, tabla detalle",
            "d2_avance_municipal, municipios",
            "Anio, departamento, provincia, municipalidad, categoria",
        ],
        [
            "03 Ranking territorial",
            "Que departamentos y provincias concentran la recaudacion?",
            "Recaudacion; municipalidades; pct ejecucion territorial",
            "Barras por departamento, tabla provincia/distrito y segmentadores geograficos",
            "d3_ranking_territorial, municipios",
            "Anio, departamento, provincia, categoria",
        ],
        [
            "04 Clasificador de ingresos",
            "De que rubros y clasificadores viene la recaudacion municipal?",
            "Recaudacion por clasificador; participacion; PIM; PIA",
            "Barras por rubro/generica/especifica, matriz drill-down, tabla top clasificadores",
            "d4_clasificador",
            "Anio, rubro, generica, especifica",
        ],
        [
            "05 Impuesto predial",
            "Como se comporta la recaudacion predial y la brecha por cobrar?",
            "Recaudacion predial total; ordinaria; coactiva; saldo predial; recuperacion",
            "Cards, barras por departamento, tendencia mensual, scatter saldo vs recaudacion",
            "d5_predial",
            "Anio estadistico, mes, departamento, tipo meta, categoria",
        ],
        [
            "06 Priorizacion municipal",
            "A que municipalidades conviene priorizar por brecha y bajo cumplimiento?",
            "Saldo predial; base imponible; recuperacion; software tributario; estado SISMEPRE; clasificacion",
            "Tabla de priorizacion, slicers por rango, barras por clasificacion, software y estado",
            "d6_priorizacion, d6_sismepre, d_renamu_software, mart_dashboard",
            "Anio, departamento, clasificacion, estado, categoria",
        ],
    ]
    dax_measures = [
        ["Recaudacion", "SUM('fact_ingresos_mensuales'[MONTO_RECAUDADO])"],
        ["PIM", "SUM('fact_ingresos_mensuales'[MONTO_PIM])"],
        ["PIA", "SUM('fact_ingresos_mensuales'[MONTO_PIA])"],
        ["Variacion PIM PIA", "SUM('fact_ingresos_mensuales'[variacion_pim_pia])"],
        ["Pct Ejecucion", "DIVIDE([Recaudacion], [PIM])"],
        ["Recaudacion Clasificador", "SUM('fact_ingresos_clasificador'[MONTO_RECAUDADO])"],
        ["Recaudacion Predial", "SUM('fact_predial_mensual'[MON_RECAUDACION_TOTAL])"],
        ["Predial Ordinario", "SUM('fact_predial_mensual'[MON_RECAUDACTUAL_ORDIN])"],
        ["Predial Coactivo", "SUM('fact_predial_mensual'[MON_RECAUDACTUAL_COAC])"],
        ["Saldo Predial", "SUM('fact_predial_mensual'[MON_SALDO_PREDIAL_TOTAL])"],
        ["Pct Recuperacion Predial", "DIVIDE([Recaudacion Predial], [Recaudacion Predial] + [Saldo Predial])"],
        ["Municipalidades", "DISTINCTCOUNT('municipios'[SEC_EJEC])"],
        ["Municipalidades SISMEPRE", "CALCULATE([Municipalidades], 'municipios'[has_sismepre] = TRUE())"],
        ["Municipalidades con Categoria", "CALCULATE([Municipalidades], NOT ISBLANK('municipios'[categoria_municipalidad]))"],
        ["Personal Municipal", "SUM('fact_renamu_gestion_tributaria'[personal_municipal_total])"],
        ["Recaudacion por Personal", "DIVIDE([Recaudacion], [Personal Municipal])"],
        ["Municipalidades con SRTM", "CALCULATE(DISTINCTCOUNT('fact_renamu_software_at'[SEC_EJEC]), 'fact_renamu_software_at'[usa_srtm_estado] = TRUE())"],
        ["Municipalidades con Software Rentas", "CALCULATE(DISTINCTCOUNT('fact_renamu_software_at'[SEC_EJEC]), 'fact_renamu_software_at'[usa_software_rentas_at] = TRUE())"],
        ["Municipalidades con Software Catastro", "CALCULATE(DISTINCTCOUNT('fact_renamu_software_at'[SEC_EJEC]), 'fact_renamu_software_at'[usa_software_catastro] = TRUE())"],
        ["Municipalidades con algun software AT", "CALCULATE(DISTINCTCOUNT('fact_renamu_software_at'[SEC_EJEC]), 'fact_renamu_software_at'[usa_al_menos_un_software_at] = TRUE())"],
        ["Brecha Predial", "[Saldo Predial]"],
    ]
    model_relationships = [
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_ingresos_mensuales", "SEC_EJEC", "1:*"],
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_ingresos_clasificador", "SEC_EJEC", "1:*"],
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_predial_mensual", "SEC_EJEC", "1:*"],
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_sismepre_cumplimiento", "SEC_EJEC", "1:*"],
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_renamu_gestion_tributaria", "SEC_EJEC", "1:*"],
        ["dim_municipalidad_gold", "SEC_EJEC", "fact_renamu_software_at", "SEC_EJEC", "1:*"],
        ["dim_ubigeo", "ubigeo_id", "dim_municipalidad_gold", "UBIGEO", "1:*"],
        ["dim_tiempo", "periodo_id", "fact_ingresos_mensuales", "periodo_id", "1:*"],
        ["dim_tiempo", "periodo_id", "fact_ingresos_clasificador", "periodo_id", "1:*"],
        ["dim_tiempo", "periodo_id", "fact_predial_mensual", "periodo_id", "1:*"],
        ["dim_clasificador_ingreso", "clasificador_id", "fact_ingresos_clasificador", "clasificador_id", "1:*"],
        ["dim_estado_sismepre", "estado_sismepre_id", "fact_sismepre_cumplimiento", "estado_sismepre_id", "1:*"],
        ["dim_formulario_sismepre", "FORMULARIO_ID", "fact_sismepre_respuestas_resumen", "FORMULARIO_ID", "1:* compuesto con año/periodo"],
        ["dim_pregunta_sismepre", "PREGUNTA_ID", "fact_sismepre_respuestas_resumen", "PREGUNTA_ID", "1:* compuesto con formulario/año/periodo"],
    ]

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Workbook Power BI Municipal"])
    readme.append(["Generado", datetime.now().isoformat(timespec="seconds")])
    readme.append(["Fuente", "data/gold"])
    readme.append(["Uso", "Importar todas las tablas desde Power BI Desktop"])
    readme.append(["Alcance", "Usa in_scope_presentacion; si la plantilla del profesor esta vacia, incluye todas las municipalidades"])
    readme.append(["Mapa", "No se usa mapa externo: las fuentes no traen latitud, longitud ni geometria"])
    readme.append([])
    readme.append(["Hoja", "Dashboard"])
    readme.append(["d1_evolucion", "Evolucion mensual del presupuesto y recaudacion"])
    readme.append(["d2_avance_municipal", "Avance de ejecucion por municipalidad"])
    readme.append(["d3_ranking_territorial", "Ranking territorial por recaudacion"])
    readme.append(["d4_clasificador", "Recaudacion segun clasificador de ingreso"])
    readme.append(["d5_predial", "Indicadores de impuesto predial"])
    readme.append(["d6_priorizacion", "Priorizacion municipal por brecha predial y cumplimiento SISMEPRE"])
    readme.append(["d6_sismepre", "Tabla auxiliar para filtros y cumplimiento SISMEPRE"])
    readme.append(["d_renamu_gestion", "Gestion tributaria RENAMU: personal y necesidades AT/catastro"])
    readme.append(["d_renamu_software", "Software tributario RENAMU: SRTM, rentas y catastro"])
    readme.append(["mart_dashboard", "Vista integrada para los 6 dashboards"])
    readme.append(["evidencia_calidad", "Soporte tecnico auditable; no es pagina de decision"])
    for cell in readme[1]:
        cell.font = Font(bold=True, size=14)
    readme.column_dimensions["A"].width = 26
    readme.column_dimensions["B"].width = 62

    add_sheet(workbook, "municipios", municipios_scope.orderBy("SEC_EJEC"))
    add_sheet(workbook, "dim_tiempo", tiempo.orderBy("periodo_id"))
    add_sheet(workbook, "dim_ubigeo", ubigeo.orderBy("ubigeo_id"))
    add_sheet(workbook, "dim_clasificador_ingreso", clasificador_dim.orderBy("clasificador_id"))
    add_sheet(workbook, "dim_estado_sismepre", estado_sismepre.orderBy("estado_sismepre_id"))
    add_sheet(workbook, "dim_formulario_sismepre", formulario_sismepre)
    add_sheet(workbook, "dim_pregunta_sismepre", pregunta_sismepre)
    add_sheet(workbook, "fact_ingresos_mensuales", ingresos)
    add_sheet(workbook, "fact_predial_mensual", predial)
    add_sheet(workbook, "fact_sismepre_cumplimiento", cumplimiento)
    add_sheet(workbook, "fact_renamu_gestion_tributaria", gestion)
    add_sheet(workbook, "fact_renamu_software_at", software)
    add_sheet(workbook, "d1_evolucion", evolucion)
    add_sheet(workbook, "d2_avance_municipal", avance)
    add_sheet(workbook, "d3_ranking_territorial", ranking)
    add_sheet(workbook, "d4_clasificador", clasificador_view)
    add_sheet(workbook, "d5_predial", predial_view)
    add_sheet(workbook, "d6_priorizacion", priorizacion_view)
    add_sheet(workbook, "d6_sismepre", sismepre_view)
    add_sheet(workbook, "d_renamu_gestion", gestion_view)
    add_sheet(workbook, "d_renamu_software", software_view)
    add_sheet(workbook, "mart_dashboard", mart_view)
    add_sheet(workbook, "evidencia_calidad", calidad_view)
    add_sheet(workbook, "dashboard_diseno", spark.createDataFrame(
        dashboard_design,
        ["pagina", "pregunta_decision", "kpis", "visuales", "hojas_fuente", "filtros"],
    ))
    add_sheet(workbook, "medidas_dax", spark.createDataFrame(dax_measures, ["medida", "formula_dax"]))
    add_sheet(workbook, "modelo_relaciones", spark.createDataFrame(
        model_relationships,
        ["tabla_origen", "columna_origen", "tabla_destino", "columna_destino", "cardinalidad"],
    ))

    saved_output = save_workbook_with_fallback(workbook, OUTPUT)
    write_dashboard_html({
        "colors": {"green2": "#45b80f"},
        "d1": collect_rows(evolucion, 2000),
        "avance": collect_rows(avance, 6000),
        "clasificador": collect_rows(clasificador_view, 6000),
        "predial": collect_rows(predial_view, 6000),
        "priorizacion": collect_rows(priorizacion_view, 6000),
        "software": collect_rows(software_view, 6000),
        "mart": collect_rows(mart_view, 8000),
    })
    print(f"Workbook created: {saved_output}")
    print(f"Dashboard HTML created: {HTML_OUTPUT}")
    print(f"Size bytes: {saved_output.stat().st_size}")
    spark.stop()


if __name__ == "__main__":
    main()
